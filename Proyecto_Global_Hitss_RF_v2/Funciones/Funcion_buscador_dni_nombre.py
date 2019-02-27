import openpyxl
from openpyxl.utils import get_column_letter
import operator
from datetime import *


def extractor_excel(excel=None,dni=None):

   book = openpyxl.load_workbook('Personal Activo HTSS Perú.xlsx')
   sheet=book.active
   nombre_completo=sheet['B7']
   print(a.value)
   #sheet=book.get_sheet_by_name('Sheet1')
   #sheet=book.get_sheet_by_name('Sample')
   #ws = book.active
   #print(sheet.cell(row=8,column=1).value)
   #sheet = book.get_sheet_by_name(book.sheetnames[0])

def extractor_nombre(cadena=None):
   nombre=""
   i=0
   for caracter in cadena:
      if caracter==" ":
         i=i+1
      if i==2:
         nombre=nombre+caracter  
   return nombre

print(extractor_nombre("VERGARA VICUÑA MARLON ANTONIO"))
extractor_excel()