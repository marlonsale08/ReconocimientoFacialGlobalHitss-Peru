import openpyxl
from openpyxl.utils import get_column_letter
import operator
from datetime import *

def sort_table(table, col=0):
    return sorted(table, key=operator.itemgetter(col))

def get_hours(table, start, end):
    data = []
    for row in table:
        if end >= row[1] and start <= row[1]:
            data.append(row)
    return data

def main(**kwargs):
    book = openpyxl.load_workbook('example.xlsx')
    book.create_sheet('Sample')
    ws = book.active
    sheet = book.get_sheet_by_name(book.sheetnames[0])
    r = sheet.max_row
    c = sheet.max_column
    rmin = sheet.min_row
    cmin = sheet.min_column
    rmatrix = 1
    cmatrix = 1
    id_cell = 0
    for i in range(rmin, r+1):
        for j in range(cmin, c+1):
            if sheet.cell(row=i, column=j).value == 'ID' and id_cell == 0:
                id_cell = j
            if sheet.cell(row=i, column=j).value == attribute_name:
                rmatrix = i
                cmatrix = j
                i = r+1
    data = []
    aux = 0
    for row in ws.iter_rows(min_row=rmatrix+1, max_col=c, max_row=r):
        if aux != str(row[cmatrix - cmin + 1].value) and aux != 0:
            break
        if str(row[cmatrix - cmin + 1].value) == search_date :
            aux = search_date
            copy_row = []
            for cell in row:
                copy_row.append(cell.value)
            data.append(copy_row)
    book = openpyxl.Workbook()
    sheet = book.active
    for row in sort_table(get_hours(data, start_hour, end_hour), cmatrix - cmin):
        sheet.append(row)
    book.save("filtered.xlsx")

if __name__ == '__main__':
    attribute_name = 'CREATE_DATE'
    search_date = '1992-12-31 00:00:00'
    start_hour = time(2, 0)
    end_hour = time(3, 30)
    kwargs = {
        "attribute_name": attribute_name,
        "search_date": search_date,
        "start_hour": start_hour,
        "end_hour": end_hour
    }

    main(**kwargs)